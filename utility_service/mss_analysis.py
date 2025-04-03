from openpyxl import load_workbook

# Load the workbook with data_only=True to extract formula results instead of formulas
file_path = "F:\\Accounts Payable\\Utilities\\Marin Sanitary Service\\2025\\test.xlsx"
wb = load_workbook(file_path, data_only=True)  # Ensures formulas return calculated values

# Load relevant sheets
source_ws = wb["Monthly Sorted by Prop"]
analysis_ws = wb["Analysis"]

# Mapping: Property Numbers (Column B in source) → Row Numbers in "Analysis" Sheet
property_mapping = {
    "10": 5,   # BON AIR ONE CO.
    "20": 6,   # BON AIR TWO CO.
    "30": 7,   # BON AIR THREE CO.
    "40": 8,   # BON AIR FOUR CO.
    "45": 9,   # HUMMINGBIRD HILL, L.P.
    "50": 10,  # BON AIR FIVE CO.
    "60": 11,  # BON AIR SIX CO., L.P.
    "70": 12,  # BON AIR SEVEN CO.
    "80": 13,  # BELARDO CO., L.P.
    "81": 14,  # BEL EAST, L.P.
    "82": 15,  # BEL WEST, L.P.
    "90": 16,  # GREENBRAE CO.
    "92": 17,  # ELISEO APARTMENTS, L.P.
    "200": 18, # MILLBRAE HIGHLANDS CO.
    "1301": 19, # ELISEO OAKS, L.P.
    "1300": 20, # ELISEO OAKS, L.P.
    "611": 21  # SBCRE, L.P.
}

# Dictionaries to store sums by property for Column F and Column S
property_sums_F = {key: 0 for key in property_mapping.keys()}  # Column F → Column E in Analysis
property_sums_S = {key: 0 for key in property_mapping.keys()}  # Column S → Column F in Analysis

# Iterate over source sheet (starting from row 4)
for row in range(4, source_ws.max_row + 1):
    property_value = str(source_ws.cell(row=row, column=2).value)  # Column B (Property)
    amount_F = source_ws.cell(row=row, column=6).value  # Column F (Total Due)
    amount_S = source_ws.cell(row=row, column=19).value  # Column S (New Data)

    if property_value in property_mapping:
        # Process Column F (Total Due)
        if amount_F is not None:
            amount_F = float(str(amount_F).replace(",", ""))  # Convert amount to float
            property_sums_F[property_value] += amount_F

        # Process Column S (New Data)
        if amount_S is not None:
            amount_S = float(str(amount_S).replace(",", ""))  # Convert amount to float
            property_sums_S[property_value] += amount_S

# Reload workbook in edit mode (without data_only=True) to write values
wb = load_workbook(file_path)
analysis_ws = wb["Analysis"]

# Write the summed values to "Analysis" sheet
for prop_num in property_mapping.keys():
    row_number = property_mapping[prop_num]
    analysis_ws.cell(row=row_number, column=5, value=property_sums_F[prop_num])  # Column E for sums from Column F
    analysis_ws.cell(row=row_number, column=6, value=property_sums_S[prop_num])  # Column F for sums from Column S

# Copy special values from "Monthly Sorted by Prop" 
special_transfers = {
    70: 23,  # Row 70 → Row 23 in Analysis
    71: 24,  # Row 71 → Row 24 in Analysis
    72: 25   # Row 72 → Row 25 in Analysis
}

for src_row, dest_row in special_transfers.items():
    # Transfer F70, F71, F72 → E23, E24, E25
    value_F = source_ws.cell(row=src_row, column=6).value  # Read from Column F
    if value_F is not None:
        analysis_ws.cell(row=dest_row, column=5, value=value_F)  # Write to Column E

    # Transfer S70, S71, S72 → F23, F24, F25
    value_S = source_ws.cell(row=src_row, column=19).value  # Read from Column S
    if value_S is not None:
        analysis_ws.cell(row=dest_row, column=6, value=value_S)  # Write to Column F

# Save the workbook
wb.save(file_path)
print("✅ Data successfully updated in 'Analysis' sheet with both special transfers and summed values.")
