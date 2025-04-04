import fitz  # PyMuPDF for PDF text extraction
import re  # Regular expressions for pattern matching
from openpyxl import load_workbook  # For handling Excel files
import os

# Define paths
pdf_folder = r"path/to/your/input/folder"
excel_file = r"path/to/your/input/folder/test.xlsx"
sheet_name = "Monthly Sorted by Prop"  # Updated sheet name

# List of account numbers in the required order - placeholder account numbers
pdf_order = [
    "10001", "10002", "10003", "10004", "10005", "10006", "10007", "10008",
    "10009", "10010", "10011", "10012", "10013", "10014", "10015", "20001",
    "20002", "20003", "20004", "20005", "20006", "20007", "20008", "20009",
    "30001", "30002", "30003", "30004", "30005", "30006", "30007", "400001",
    "400002", "5001", "30008", "30009", "30010", "30011", "6001", "6002",
    "6003", "6004", "6005", "6006", "6007", "6008", "6009", "6010", "6011",
    "6012", "6013", "6014", "6015", "6016", "6017", "6018", "6019", "6020",
    "7001", "7002", "80001", "80002", "6021", "6022", "900001", "900002",
    "90001", "6023"
]

# Get all PDF files in the folder
pdf_files = os.listdir(pdf_folder)

# Filter and sort PDFs based on account numbers
sorted_pdfs = []
for account_number in pdf_order:
    matched_files = [file for file in pdf_files if account_number in file and file.endswith(".pdf")]
    sorted_pdfs.extend(matched_files)  # Maintain order from pdf_order

# Compile a regex pattern to extract "Total Due" amounts (handles numbers with or without commas)
amount_pattern = re.compile(r"TOTAL\s+DUE\s+(\d{1,3}(?:,\d{3})*\.\d{2}|\d+\.\d{2})", re.IGNORECASE)

# Open the Excel workbook
wb = load_workbook(excel_file)
ws = wb[sheet_name]

# Start writing data to Excel from row 4 onwards
current_row = 4

# Iterate through the sorted PDFs
for pdf_file in sorted_pdfs:
    pdf_path = os.path.join(pdf_folder, pdf_file)

    doc = fitz.open(pdf_path)
    text = ""

    # Extract text from each page
    for page in doc:
        text += page.get_text()
    doc.close()

    # Debugging: Print the extracted text
    # print(f"Extracted text from {pdf_file}:\n{text}\n{'-'*50}")

    # Search for the 'Total Due' amount
    match = amount_pattern.search(text)
    if match:
        amount_due = match.group(1)  # Extract the amount
        print(f':D Found amount in {pdf_file}: {amount_due}')
        
        # Special case for accounts 27922 and 2494
        if "27922" in pdf_file:
            ws.cell(row=72, column=6, value=amount_due)  # Row 72, Column F (Excel column 6)
        elif "2494" in pdf_file:
            ws.cell(row=75, column=6, value=amount_due)  # Row 75, Column F (Excel column 6)
        else:
            ws.cell(row=current_row, column=6, value=amount_due)  # Insert into next available row

    else:
        print(f'No amount found in {pdf_file}')

    # Move to the next row, except for special cases
    if "27922" not in pdf_file and "2494" not in pdf_file:
        current_row += 1
  

# Save the updated Excel file
wb.save("path/to/your/input/folder/test.xlsx")
print(":D Data extraction and insertion completed successfully.")
