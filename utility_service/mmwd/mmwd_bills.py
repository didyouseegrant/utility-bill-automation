import fitz  # PyMuPDF for PDF text extraction
import re  # Regular expressions for pattern matching
from openpyxl import load_workbook  # For handling Excel files
import os

# Define paths
pdf_folder = r"C:\Accounts Payable\Utilites\MMWD\Bills"
excel_file = r"C:\Accounts Payable\Utilites\MMWD\test.xlsx"
sheet_name = "Monthly Comparison"  # Updated sheet name

# List of account numbers in the required order
pdf_order = [
    111111, 222222, 333333, 444444, 555555, 666666, 777777, 888888, 999999, 
    101111, 112222, 123333, 134444, 145555, 156666, 167777, 178888, 189999, 
    191111, 202222, 213333, 224444, 235555, 246666, 257777, 268888, 279999, 
    281111, 292222, 303333, 314444, 325555, 336666, 347777, 358888, 369999, 
    371111, 382222, 393333, 404444, 415555, 426666, 437777, 448888, 459999, 
    461111, 472222, 483333, 494444, 505555, 516666
]

# Get all PDF files in the folder
pdf_files = os.listdir(pdf_folder)

# Filter and sort PDFs based on account numbers
sorted_pdfs = []
for account_number in pdf_order:
    account_str = str(account_number)  # Convert account number to string
    matched_files = [file for file in pdf_files if file.startswith(account_str) and file.endswith(".pdf")]
    sorted_pdfs.extend(matched_files)  # Maintain order from pdf_order

# Compile a regex pattern to extract "Total Amount Due" amounts (handles numbers with or without commas)
amount_pattern = re.compile(r"TOTAL\s+AMOUNT\s+DUE\s+(\d{1,3}(?:,\d{3})*\.\d{2}|\d+\.\d{2})", re.IGNORECASE)
gallons_pattern = re.compile(r"GALLONS\s+([\d,]+)", re.IGNORECASE)  # Matches "GALLONS" followed by a number

# Open the Excel workbook
wb = load_workbook(excel_file)
ws = wb[sheet_name]

# Start writing data to Excel from row 4 onwards
current_row = 4

# Iterate through the sorted PDFs
for pdf_file in sorted_pdfs:
    pdf_path = os.path.join(pdf_folder, pdf_file)

    doc = fitz.open(pdf_path)
    text = ""

    # Extract text from each page
    for page in doc:
        text += page.get_text()
    doc.close()

    # Debugging: Print the extracted text
    # print(f"Extracted text from {pdf_file}:\n{text}\n{'-'*50}")

    # Extract 'Total Amount Due'
    amount_match = amount_pattern.search(text)
    amount_due = amount_match.group(1) if amount_match else None

    # Extract first 'Gallons' value
    gallons_match = gallons_pattern.search(text)
    gallons_used = gallons_match.group(1).replace(",", "") if gallons_match else None  # Remove commas

    # Debugging output
    print(f":D Found values in {pdf_file}: Amount Due = {amount_due}, Gallons = {gallons_used}")

    # Insert values into Excel
    if amount_due:
        ws.cell(row=current_row, column=7, value=amount_due)  # Column G (Excel column 7)
    if gallons_used:
        ws.cell(row=current_row, column=6, value=gallons_used)  # Column F (Excel column 6)

    # Move to the next row for the next PDF
    current_row += 1
  

# Save the updated Excel file
wb.save("C:\\Accounts Payable\\Utilites\\MMWD\\test2.xlsx")
print(":3 Data extraction and insertion completed successfully.")
