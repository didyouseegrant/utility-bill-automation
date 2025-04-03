from openpyxl import load_workbook

# Load the workbook with data_only=True to extract formula results instead of formulas
file_path = "F:\\Accounts Payable\\Utilities\\Marin Municipal Water\\2025\\test.xlsx"
wb = load_workbook(file_path, data_only=True)  # Ensures formulas return calculated values

# Load relevant sheets
source_ws = wb["Monthly Comparison"]
analysis_ws = wb["Analysis"]

# Mapping: Property Numbers (Column D in source) → Row Numbers in "Analysis" Sheet
property_mapping = {
    "611": 4,  # SBCRE, L.P.
    "5": 5,   # GMI (YARD)
    "5CAM": 6,   # CAM
    "30/40/50": 7,   # BON AIR THREE CO.
    "30/40/50": 8,   # BON AIR FOUR CO.
    "45": 9,   # HUMMINGBIRD HILL, L.P.
    "30/40/50": 10,  # BON AIR FIVE CO.
    "60": 11,  # BON AIR SIX CO., L.P.
    "70": 12,  # BON AIR SEVEN CO.
    "80": 13,  # BELARDO CO., L.P (B52-B55)
    "80P": 14, # BELARDO CO., L.P (80POOL2)
    "81": 15,  # BEL EAST, L.P.
    "82": 16,  # BEL WEST, L.P.
    "82P": 17, # BEL WEST. L.P. (82POOL3)
    "90": 18,  # GREENBRAE CO.
    "92": 19,  # ELISEO APARTMENTS, L.P.
    "200": 20, # MILLBRAE HIGHLANDS CO.
    "SIC50": 21, # SICRE, INC. (50BAC)
    "214P": 22, # SICRE, INC. (214POOL)
    "222": 23, # BON AIR, L.P.
    "BAC350": 24, # BON AIR DEVELOPMENT, L.P.
    "1300": 25, # ELISEO OAKS, L.P.
    "1301": 26, # ELISEO OAKS, L.P.
    "501": 27 # 501 SFD
}

# Dictionaries to store sums by property for Column F and Column S
property_sums_G = {key: 0 for key in property_mapping.keys()}  # Column G → Column B in Analysis 
property_sums_F = {key: 0 for key in property_mapping.keys()}  # Column F → Column F in Analysis
property_sums_S = {key: 0 for key in property_mapping.keys()}  # Column S → Column C in Analysis
property_sums_R = {key: 0 for key in property_mapping.keys()}  # Column R → Column G in Analysis

# Iterate over source sheet (starting from row 4)
for row in range(4, source_ws.max_row + 1):
    property_value = str(source_ws.cell(row=row, column=4).value)  # Column D (Property)
    amount_G = source_ws.cell(row=row, column=7).value  # Column G (Total Due)
    amount_F = source_ws.cell(row=row, column=6).value  # Column F (Total Gallons)
    amount_S = source_ws.cell(row=row, column=19).value  # Column S (Last Year Due)
    amount_R = source_ws.cell(row=row, column=18).value  # Column R (Last Year Gallons)

    if property_value in property_mapping:
        # Process Column G (Total Due)
        if amount_G is not None:
            amount_G = float(str(amount_G).replace(",", ""))  # Convert amount to float
            property_sums_G[property_value] += amount_G

        # Process Column F (Total Gallons)
        if amount_F is not None:
            amount_F = float(str(amount_F).replace(",", ""))  # Convert amount to float
            property_sums_F[property_value] += amount_F

        # Process Column S (Last Year Due)
        if amount_S is not None:
            amount_S = float(str(amount_S).replace(",", ""))  # Convert amount to float
            property_sums_S[property_value] += amount_S

        # Process Column R (Last Year Gallons)
        if amount_R is not None:
            amount_R = float(str(amount_R).replace(",", ""))  # Convert amount to float
            property_sums_R[property_value] += amount_R

# Reload workbook in edit mode (without data_only=True) to write values
wb = load_workbook(file_path)
analysis_ws = wb["Analysis"]

# Write the summed values to "Analysis" sheet
for prop_num in property_mapping.keys():
    row_number = property_mapping[prop_num]
    analysis_ws.cell(row=row_number, column=2, value=property_sums_G[prop_num])  # Column B for sums from Column G
    analysis_ws.cell(row=row_number, column=6, value=property_sums_F[prop_num])  # Column F for sums from Column F
    analysis_ws.cell(row=row_number, column=3, value=property_sums_S[prop_num])  # Column C for sums from Column S
    analysis_ws.cell(row=row_number, column=7, value=property_sums_R[prop_num])  # Column G for sums from Column R



# Save the workbook
wb.save(file_path)
print("✅ Data successfully updated in 'Analysis' sheet with both special transfers and summed values.")
