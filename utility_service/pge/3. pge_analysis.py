from openpyxl import load_workbook

# Load the workbook with data_only=True to extract formula results instead of formulas
file_path = "path/to/your/input/folder/test.xlsx"
wb = load_workbook(file_path, data_only=True)  # Ensures formulas return calculated values

# Load relevant sheets
source_ws = wb["Monthly Comparison"]
analysis_ws = wb["Analysis"]

# Mapping: Property Numbers (Column D in source) -> Row Numbers in "Analysis" Sheet
# These are placeholders
property_mapping = {
    "10": 4,  # Property name
    "20": 5,   # Property name
    "30": 6,   # Property name
    "40": 7,   # Property name
    "45": 8,   # Property name
    "50": 9,  # Property name
    "60": 10,  # Property name
    "70": 11,  # Property name
    "80": 12,  # Property name
    "80P": 13, # Property name
    "81": 14,  # Property name
    "82": 15,  # Property name
    "82P": 16, # Property name
    "90": 17,  # Property name
    "92": 18,  # Property name
    "200": 19, # Property name
    "214P": 20, # Property name
    "S50": 21, # Property name
    "222": 22, # Property name
    "350": 23, # Property name
    "1300": 24, # Property name
    "1301": 25, # Property name
    "611": 26, # Property name
    "501": 27, # Property name
    "5B": 30, # Property name
    "5C": 31 # Property name
}

# Dictionaries to store sums by property for Column F and Column S
property_sums_F = {key: 0 for key in property_mapping.keys()}  # Column F -> Column E Current Charges in Analysis 
property_sums_G = {key: 0 for key in property_mapping.keys()}  # Column G -> Column F Check Charges in Analysis
property_sums_H = {key: 0 for key in property_mapping.keys()}  # Column H -> Column H Prior Month Charges in Analysis
property_sums_W = {key: 0 for key in property_mapping.keys()}  # Column W -> Column G Prior Year Charges in Analysis


# Iterate over source sheet (starting from row 4)
for row in range(4, source_ws.max_row + 1):
    property_value = str(source_ws.cell(row=row, column=4).value)  # Column D (Property)
    amount_F = source_ws.cell(row=row, column=6).value  # Column F (Current Charges)
    amount_G = source_ws.cell(row=row, column=7).value  # Column G (Check Charges)
    amount_H = source_ws.cell(row=row, column=8).value  # Column H (Prior Month Charges)
    amount_W = source_ws.cell(row=row, column=23).value  # Column W (Prior Year Charges)
    

    if property_value in property_mapping:
        # Process Column F (Current Charges)
        if amount_F is not None:
            amount_F = float(str(amount_F).replace(",", ""))  # Convert amount to float
            property_sums_F[property_value] += amount_F
        
        # Process Column G (Check Charges)
        if amount_G is not None:
            amount_G = float(str(amount_G).replace(",", ""))  # Convert amount to float
            property_sums_G[property_value] += amount_G

        # Process Column H (Prior Month Charges)
        if amount_H is not None:
            amount_H = float(str(amount_H).replace(",", "")) # Convert amount to float
            property_sums_H[property_value] += amount_H

        # Process Column S (Prior Year Charges)
        if amount_W is not None:
            amount_W = float(str(amount_X).replace(",", ""))  # Convert amount to float
            property_sums_W[property_value] += amount_W


# Reload workbook in edit mode (without data_only=True) to write values
wb = load_workbook(file_path)
analysis_ws = wb["Analysis"]

# Write the summed values to "Analysis" sheet
for prop_num in property_mapping.keys():
    row_number = property_mapping[prop_num]
    analysis_ws.cell(row=row_number, column=5, value=property_sums_F[prop_num])  # Column E for sums from Column F
    analysis_ws.cell(row=row_number, column=6, value=property_sums_G[prop_num])  # Column F for sums from Column G
    analysis_ws.cell(row=row_number, column=4, value=property_sums_H[prop_num])  # Column H for sums from Column H
    analysis_ws.cell(row=row_number, column=7, value=property_sums_W[prop_num])  # Column G for sums from Column W


# Save the workbook
wb.save(file_path)
print("Data successfully updated in 'Analysis' sheet with both special transfers and summed values.")
