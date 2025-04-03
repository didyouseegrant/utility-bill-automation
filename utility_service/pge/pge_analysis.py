from openpyxl import load_workbook

# Load the workbook with data_only=True to extract formula results instead of formulas
file_path = "F:\\Accounts Payable\\Utilities\\PG&E\\2025\\test.xlsx"
wb = load_workbook(file_path, data_only=True)  # Ensures formulas return calculated values

# Load relevant sheets
source_ws = wb["Monthly Comparison"]
analysis_ws = wb["Analysis"]

# Mapping: Property Numbers (Column D in source) → Row Numbers in "Analysis" Sheet
property_mapping = {
    "10": 4,  # BON AIR ONE CO.
    "20": 5,   # BON AIR TWO CO.
    "30": 6,   # BON AIR THREE CO.
    "40": 7,   # BON AIR FOUR CO.
    "45": 8,   # HUMMINGBIRD HILL, L.P.
    "50": 9,  # BON AIR FIVE CO.
    "60": 10,  # BON AIR SIX CO., L.P.
    "70": 11,  # BON AIR SEVEN CO.
    "80": 12,  # BELARDO CO., L.P (80)
    "80P": 13, # BELARDO CO., L.P (80POOL2)
    "81": 14,  # BEL EAST, L.P.
    "82": 15,  # BEL WEST, L.P.
    "82P": 16, # BEL WEST. L.P. (82POOL3)
    "90": 17,  # GREENBRAE CO.
    "92": 18,  # ELISEO APARTMENTS, L.P.
    "200": 19, # MILLBRAE HIGHLANDS CO.
    "214P": 20, # SICRE, INC. (214POOL)
    "SIC50": 21, # SICRE, INC. (50BAC)
    "222": 22, # BON AIR, L.P.
    "BAC350": 23, # BON AIR DEVELOPMENT, L.P.
    "1300": 24, # ELISEO OAKS, L.P. (1300 ELISEO)
    "1301": 25, # ELISEO OAKS, L.P. (100 1301)
    "611": 26, # SBCRE, L.P. (22B)
    "501": 27, # 501 SFD
    "5BAM": 30, # RENTAL OFFICE
    "5CAM": 31 # CAM
}

# Dictionaries to store sums by property for Column F and Column S
property_sums_F = {key: 0 for key in property_mapping.keys()}  # Column F → Column E Current Charges in Analysis 
property_sums_G = {key: 0 for key in property_mapping.keys()}  # Column G → Column F Check Charges in Analysis
property_sums_X = {key: 0 for key in property_mapping.keys()}  # Column X → Column G Prior Year Charges in Analysis


# Iterate over source sheet (starting from row 4)
for row in range(4, source_ws.max_row + 1):
    property_value = str(source_ws.cell(row=row, column=4).value)  # Column D (Property)
    amount_F = source_ws.cell(row=row, column=6).value  # Column F (Current Charges)
    amount_G = source_ws.cell(row=row, column=7).value  # Column G (Check Charges)
    amount_X = source_ws.cell(row=row, column=24).value  # Column X (Prior Year Charges)
    

    if property_value in property_mapping:
        # Process Column F (Current Charges)
        if amount_F is not None:
            amount_F = float(str(amount_F).replace(",", ""))  # Convert amount to float
            property_sums_F[property_value] += amount_F
        
        # Process Column G (Check Charges)
        if amount_G is not None:
            amount_G = float(str(amount_G).replace(",", ""))  # Convert amount to float
            property_sums_G[property_value] += amount_G

        # Process Column S (Prior Year Charges)
        if amount_X is not None:
            amount_X = float(str(amount_X).replace(",", ""))  # Convert amount to float
            property_sums_X[property_value] += amount_X


# Reload workbook in edit mode (without data_only=True) to write values
wb = load_workbook(file_path)
analysis_ws = wb["Analysis"]

# Write the summed values to "Analysis" sheet
for prop_num in property_mapping.keys():
    row_number = property_mapping[prop_num]
    analysis_ws.cell(row=row_number, column=5, value=property_sums_F[prop_num])  # Column E for sums from Column F
    analysis_ws.cell(row=row_number, column=6, value=property_sums_G[prop_num])  # Column F for sums from Column G
    analysis_ws.cell(row=row_number, column=7, value=property_sums_X[prop_num])  # Column G for sums from Column X




# Save the workbook
wb.save(file_path)
print("✅ Data successfully updated in 'Analysis' sheet with both special transfers and summed values.")
